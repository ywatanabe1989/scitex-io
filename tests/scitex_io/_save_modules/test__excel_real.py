"""Real round-trip tests for scitex_io._save_modules._excel."""


from __future__ import annotations
import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from scitex_io._save_modules._excel import (
    _apply_stats_styling,
    _generate_pval_variants,
    _is_statistical_results,
    save_excel,
)


def test_generate_pval_variants_includes_common_names():
    v = _generate_pval_variants()
    for name in ["p", "pvalue", "padj", "p_adj", "p-val"]:
        assert name in v


def test_is_statistical_results_true_for_pvalue_col():
    df = pd.DataFrame({"feature": ["a", "b"], "pvalue": [0.04, 0.6]})
    assert _is_statistical_results(df) is True


def test_is_statistical_results_false_for_non_df():
    assert _is_statistical_results({"pvalue": [0.1]}) is False


def test_is_statistical_results_false_without_pval_col():
    df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
    assert _is_statistical_results(df) is False


def test_save_excel_dataframe_round_trip(tmp_path):
    spath = str(tmp_path / "a.xlsx")
    df = pd.DataFrame({"a": [1, 2, 3], "b": ["x", "y", "z"]})
    save_excel(df, spath)
    out = pd.read_excel(spath)
    pd.testing.assert_frame_equal(out, df)


def test_save_excel_dict_round_trip(tmp_path):
    spath = str(tmp_path / "b.xlsx")
    save_excel({"a": [1, 2], "b": [3, 4]}, spath)
    out = pd.read_excel(spath)
    assert list(out.columns) == ["a", "b"]
    assert out.shape == (2, 2)


def test_save_excel_list_of_dicts(tmp_path):
    spath = str(tmp_path / "c.xlsx")
    save_excel([{"a": 1, "b": 2}, {"a": 3, "b": 4}], spath)
    out = pd.read_excel(spath)
    assert out.shape == (2, 2)


def test_save_excel_numpy_array(tmp_path):
    spath = str(tmp_path / "d.xlsx")
    arr = np.arange(6).reshape(3, 2)
    save_excel(arr, spath)
    out = pd.read_excel(spath)
    assert out.shape == (3, 2)


def test_save_excel_invalid_type_raises(tmp_path):
    spath = str(tmp_path / "e.xlsx")
    with pytest.raises(ValueError):
        save_excel(42, spath)


def test_save_excel_with_styling_applied(tmp_path):
    """Stat-results DataFrame triggers conditional formatting."""
    spath = str(tmp_path / "f.xlsx")
    df = pd.DataFrame(
        {
            "feature": ["a", "b", "c", "d"],
            "pvalue": [0.0001, 0.005, 0.04, 0.5],
        }
    )
    save_excel(df, spath, style=True)
    wb = load_workbook(spath)
    ws = wb.active
    # Header row 1 + 4 data rows
    # Find pvalue column
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    pcol = headers["pvalue"]
    # Row 2 (p=0.0001) should have a red fill
    red_cell = ws.cell(row=2, column=pcol)
    assert red_cell.fill.start_color.rgb in ("00FF6B6B", "FF6B6B", "FFFF6B6B")
    # Row 5 (p=0.5) should be gray
    gray_cell = ws.cell(row=5, column=pcol)
    assert "E8E8E8" in str(gray_cell.fill.start_color.rgb)
    # Frozen pane header
    assert ws.freeze_panes == "A2"


def test_save_excel_styling_disabled(tmp_path):
    spath = str(tmp_path / "g.xlsx")
    df = pd.DataFrame({"feature": ["a"], "pvalue": [0.001]})
    save_excel(df, spath, style=False)
    wb = load_workbook(spath)
    ws = wb.active
    # Without styling there is no freeze_panes set
    assert ws.freeze_panes is None


def test_apply_stats_styling_direct(tmp_path):
    """Exercise _apply_stats_styling directly for the p<0.01 branch."""
    spath = str(tmp_path / "h.xlsx")
    df = pd.DataFrame({"feature": ["a", "b"], "pvalue": [0.008, 0.06]})
    df.to_excel(spath, index=False)
    _apply_stats_styling(df, spath)
    wb = load_workbook(spath)
    ws = wb.active
    # The pvalue column for p=0.008 should be orange (FFA500)
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    pcol = headers["pvalue"]
    cell = ws.cell(row=2, column=pcol)
    assert "FFA500" in str(cell.fill.start_color.rgb)
