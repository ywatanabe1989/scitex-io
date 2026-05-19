"""Real round-trip tests for scitex_io._save_modules._excel."""


from __future__ import annotations
import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from scitex_io._save_modules._excel import (
    _apply_stats_styling,
    _generate_pval_variants,
    _is_statistical_results,
    save_excel,
)


def test_generate_pval_variants_includes_common_names():
    # Arrange
    # Act
    # Assert
    # Arrange
    # Act
    v = _generate_pval_variants()
    # Assert
    assert all(name in v for name in ['p', 'pvalue', 'padj', 'p_adj', 'p-val'])


def test_is_statistical_results_true_for_pvalue_col():
    # Arrange
    # Act
    # Arrange
    # Act
    df = pd.DataFrame({"feature": ["a", "b"], "pvalue": [0.04, 0.6]})
    # Assert
    # Assert
    assert _is_statistical_results(df) is True


def test_is_statistical_results_false_for_non_df():
    # Arrange
    # Act
    # Assert
    # Arrange
    # Act
    # Assert
    assert _is_statistical_results({"pvalue": [0.1]}) is False


def test_is_statistical_results_false_without_pval_col():
    # Arrange
    # Act
    # Arrange
    # Act
    df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
    # Assert
    # Assert
    assert _is_statistical_results(df) is False


def test_save_excel_dataframe_round_trip(tmp_path):
    # Arrange
    # Act
    # Assert
    # Arrange
    # Act
    # Assert
    spath = str(tmp_path / "a.xlsx")
    df = pd.DataFrame({"a": [1, 2, 3], "b": ["x", "y", "z"]})
    save_excel(df, spath)
    out = pd.read_excel(spath)
    pd.testing.assert_frame_equal(out, df)


def test_save_excel_dict_round_trip_list_out_columns_a_b(tmp_path):
    # Arrange
    # Arrange
    spath = str(tmp_path / "b.xlsx")
    save_excel({"a": [1, 2], "b": [3, 4]}, spath)
    # Act
    out = pd.read_excel(spath)
    # Act
    # Assert
    # Assert
    assert list(out.columns) == ["a", "b"]


def test_save_excel_dict_round_trip_out_shape_equals_n_2_2(tmp_path):
    # Arrange
    # Arrange
    spath = str(tmp_path / "b.xlsx")
    save_excel({"a": [1, 2], "b": [3, 4]}, spath)
    # Act
    out = pd.read_excel(spath)
    # Act
    # Assert
    # Assert
    assert out.shape == (2, 2)




def test_save_excel_list_of_dicts(tmp_path):
    # Arrange
    # Arrange
    spath = str(tmp_path / "c.xlsx")
    save_excel([{"a": 1, "b": 2}, {"a": 3, "b": 4}], spath)
    # Act
    # Act
    out = pd.read_excel(spath)
    # Assert
    # Assert
    assert out.shape == (2, 2)


def test_save_excel_numpy_array(tmp_path):
    # Arrange
    # Arrange
    spath = str(tmp_path / "d.xlsx")
    arr = np.arange(6).reshape(3, 2)
    save_excel(arr, spath)
    # Act
    # Act
    out = pd.read_excel(spath)
    # Assert
    # Assert
    assert out.shape == (3, 2)


def test_save_excel_invalid_type_raises(tmp_path):
    # Arrange
    # Act
    # Arrange
    # Act
    spath = str(tmp_path / "e.xlsx")
    # Assert
    # Assert
    with pytest.raises(ValueError):
        save_excel(42, spath)


def test_save_excel_with_styling_applied_red_cell_fill_start_color_rgb_in_00ff6b6b_ff6b6b_ffff6b6b(tmp_path):
    # Arrange
    # Arrange
    spath = str(tmp_path / "f.xlsx")
    df = pd.DataFrame(
        {
            "feature": ["a", "b", "c", "d"],
            "pvalue": [0.0001, 0.005, 0.04, 0.5],
        }
    )
    save_excel(df, spath, style=True)
    wb = load_workbook(spath)
    ws = wb.active
    # Header row 1 + 4 data rows
    # Find pvalue column
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    pcol = headers["pvalue"]
    # Row 2 (p=0.0001) should have a red fill
    # Act
    red_cell = ws.cell(row=2, column=pcol)
    # Act
    # Assert
    # Assert
    assert red_cell.fill.start_color.rgb in ("00FF6B6B", "FF6B6B", "FFFF6B6B")


def test_save_excel_with_styling_applied_e8e8e8_in_str_gray_cell_fill_start_color_rgb(tmp_path):
    # Arrange
    # Arrange
    spath = str(tmp_path / "f.xlsx")
    df = pd.DataFrame(
        {
            "feature": ["a", "b", "c", "d"],
            "pvalue": [0.0001, 0.005, 0.04, 0.5],
        }
    )
    save_excel(df, spath, style=True)
    wb = load_workbook(spath)
    ws = wb.active
    # Header row 1 + 4 data rows
    # Find pvalue column
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    pcol = headers["pvalue"]
    # Row 2 (p=0.0001) should have a red fill
    # Act
    red_cell = ws.cell(row=2, column=pcol)
    # Assert
    assert red_cell.fill.start_color.rgb in ("00FF6B6B", "FF6B6B", "FFFF6B6B")
    # Row 5 (p=0.5) should be gray
    gray_cell = ws.cell(row=5, column=pcol)
    # Act
    # Assert
    assert "E8E8E8" in str(gray_cell.fill.start_color.rgb)


def test_save_excel_with_styling_applied_ws_freeze_panes_equals_a2(tmp_path):
    # Arrange
    # Arrange
    spath = str(tmp_path / "f.xlsx")
    df = pd.DataFrame(
        {
            "feature": ["a", "b", "c", "d"],
            "pvalue": [0.0001, 0.005, 0.04, 0.5],
        }
    )
    save_excel(df, spath, style=True)
    wb = load_workbook(spath)
    ws = wb.active
    # Header row 1 + 4 data rows
    # Find pvalue column
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    pcol = headers["pvalue"]
    # Row 2 (p=0.0001) should have a red fill
    # Act
    red_cell = ws.cell(row=2, column=pcol)
    # Assert
    assert red_cell.fill.start_color.rgb in ("00FF6B6B", "FF6B6B", "FFFF6B6B")
    # Row 5 (p=0.5) should be gray
    gray_cell = ws.cell(row=5, column=pcol)
    # Act
    # Assert
    assert ws.freeze_panes == "A2"




def test_save_excel_styling_disabled(tmp_path):
    # Arrange
    # Arrange
    spath = str(tmp_path / "g.xlsx")
    df = pd.DataFrame({"feature": ["a"], "pvalue": [0.001]})
    save_excel(df, spath, style=False)
    wb = load_workbook(spath)
    # Act
    # Act
    ws = wb.active
    # Without styling there is no freeze_panes set
    # Assert
    # Assert
    assert ws.freeze_panes is None


def test_apply_stats_styling_direct(tmp_path):
    """Exercise _apply_stats_styling directly for the p<0.01 branch."""
    # Arrange
    spath = str(tmp_path / "h.xlsx")
    df = pd.DataFrame({"feature": ["a", "b"], "pvalue": [0.008, 0.06]})
    df.to_excel(spath, index=False)
    _apply_stats_styling(df, spath)
    wb = load_workbook(spath)
    ws = wb.active
    # The pvalue column for p=0.008 should be orange (FFA500)
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    pcol = headers["pvalue"]
    # Act
    cell = ws.cell(row=2, column=pcol)
    # Assert
    assert "FFA500" in str(cell.fill.start_color.rgb)
